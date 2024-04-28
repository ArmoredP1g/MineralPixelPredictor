import sys
sys.path.append('./')
from unicodedata import name
from models.models import Conv_Diff
import json
import h5py
import ast
import numpy as np
import spectral
from spectral import imshow
from PIL import Image
spectral.settings.envi_support_nonlowercase_params = True
from dataloaders.dataloaders import dataset_iron_balanced_mixed
import torch
from torch.utils.data import DataLoader
from torch.utils.data  import random_split
from trainer import train_regression
from configs.training_cfg import *
import torch
import os
import openpyxl

torch.autograd.set_detect_anomaly(True)
pool = torch.nn.AvgPool2d(3,3)
mask_rgb_values = [[255,242,0],[34,177,76],[255,0,88]]

def test_data(list):
    test_data_list = []
    for id in list:
        pixel_list = []
        imgid, sampleid = id.split('_')
        sampleid = ord(sampleid) - 65
        img_data = spectral.envi.open(dataset_path+"/spectral_data/{}-Radiance From Raw Data-Reflectance from Radiance Data and Measured Reference Spectrum.bip.hdr".format(imgid))
        gt = ast.literal_eval(img_data.metadata['gt_TFe'])
        img_data = torch.Tensor(img_data.asarray()/6000)[:,:,:]
        img_data = pool(img_data.permute(2,0,1)).permute(1,2,0)
        mask = np.array(Image.open(dataset_path+"/spectral_data/{}-Radiance From Raw Data-Reflectance from Radiance Data and Measured Reference Spectrum.bip.hdr_mask.png".format(imgid)))
        row, col, _ = img_data.shape
        for r in range(row):
            for c in range(col):
                if mask[r*3+1,c*3+1].tolist() == mask_rgb_values[sampleid]:
                    pixel_list.append(img_data[r,c].unsqueeze(0))

        pixel_list = torch.cat(pixel_list, dim=0)
        test_data_list.append({
            "tensor": pixel_list.to(device),
            "gt": torch.Tensor([gt[sampleid]]).to(device)
        })
    return test_data_list



if __name__ == "__main__":
    with torch.no_grad():
        # all dataset
        # all_samples = ['13_A','11_A','12_A','12_C','25_A','42_B','55_C','15_A',
        #             '56_B','4_B','42_A','57_A','14_B','36_B','43_C','26_A',
        #             '9_C','43_A','53_A','3_B','30_C','27_A','22_B','27_C','31_C',
        #             '53_B','32_A','6_B','52_B','8_B','41_B','34_A',
        #             '7_B','53_C','54_C','29_B','16_B','47_A','49_B','10_C',
        #             '21_C','31_B','50_A','18_A','22_C','52_C','38_A','17_A',
        #             '59_A','4_A','57_B','33_C','7_A','49_C','58_B','4_C',
        #             '52_A','17_C','23_A','7_C','46_B','30_B','46_A','18_C',
        #             '24_A','55_A','40_A','55_B','6_A','59_B','3_C','27_B',
        #             '18_B','5_A','29_A','25_B','49_A','32_C','45_C','12_B',
        #             '20_A','9_A','28_C','29_C','5_C','46_C','14_C','19_A',
        #             '23_B','9_B','40_B','35_C','13_C','50_B','35_B','15_B',
        #             '44_C','45_A','23_C','1_C','1_B','35_A','32_B','6_C',
        #             '51_B','28_B','2_B','58_C','38_C','2_A','26_B','2_C',
        #             '16_C','43_B','24_C','54_B','15_C','42_C','36_A','37_A',
        #             '41_C','44_B','19_C','51_C','1_A','39_B','28_A','58_A',
        #             '39_C','30_A','39_A','54_A','61_C','61_A','37_B','48_C',
        #             '21_A','22_A','48_B','48_A'] + \
        #             ['14_A',
        #             '11_C','16_A','13_B','21_B','36_C','31_A','34_B','56_A',
        #             '47_C','8_A','19_B','3_A','62_A','33_B','10_A','24_B','60_B',
        #             '11_B','59_C','10_B','60_C','45_B','8_C','41_A','38_B',
        #             '57_C','50_C','61_B','37_C']
        all_samples = ['13_A','11_A','12_A','12_C','25_A','42_B','55_C','15_A',
                    '56_B','4_B','42_A','57_A','14_B','36_B','43_C','26_A',
                    '9_C','43_A','53_A','3_B','30_C','27_A','22_B','27_C','31_C',
                    '53_B','32_A','6_B','52_B','8_B','41_B','31_A','34_A',
                    '7_B','53_C','54_C','29_B','16_B','47_A','49_B','10_C',
                    '21_C','50_A','18_A','22_C','52_C','38_A','17_A',
                    '59_A','4_A','57_B','33_C','7_A','49_C','58_B','4_C',
                    '52_A','17_C','23_A','7_C','46_B','30_B','46_A','18_C',
                    '24_A','55_A','40_A','55_B','6_A','59_B','3_C','27_B',
                    '18_B','5_A','29_A','25_B','49_A','32_C','45_C','12_B',
                    '20_A','9_A','28_C','29_C','5_C','46_C','14_C','19_A',
                    '23_B','9_B','40_B','35_C','13_C','50_B','35_B','15_B',
                    '45_A','23_C','1_C','1_B','35_A','32_B','6_C',
                    '51_B','28_B','2_B','58_C','38_C','2_A','26_B','2_C',
                    '16_C','43_B','24_C','54_B','15_C','42_C','36_A','37_A',
                    '41_C','44_B','19_C','51_C','1_A','39_B','28_A',
                    '39_C','30_A','39_A','54_A','61_C','61_A','37_B','48_C',
                    '21_A','22_A','48_B','48_A'] + \
                    ['14_A',
                    '11_C','16_A','13_B','21_B','36_C','34_B','56_A',
                    '47_C','8_A','19_B','3_A','62_A','33_B','10_A','24_B','60_B',
                    '11_B','59_C','10_B','8_C','41_A','38_B',
                    '57_C','61_B','37_C']
        
        pt_path_list = ['D:\\source\\repos\\Pixel-wise-hyperspectral-feature-classification-experiment\\ckpt\\1e-4_0.93_1000_1e-7\\fold0_step100000.pt',
                        'D:\\source\\repos\\Pixel-wise-hyperspectral-feature-classification-experiment\\ckpt\\1e-4_0.93_1000_1e-7\\fold1_step100000.pt',
                        'D:\\source\\repos\\Pixel-wise-hyperspectral-feature-classification-experiment\\ckpt\\1e-4_0.93_1000_1e-7\\fold2_step100000.pt',
                        'D:\\source\\repos\\Pixel-wise-hyperspectral-feature-classification-experiment\\ckpt\\1e-4_0.93_1000_1e-7\\fold3_step100000.pt',
                        'D:\\source\\repos\\Pixel-wise-hyperspectral-feature-classification-experiment\\ckpt\\1e-4_0.93_1000_1e-7\\fold4_step100000.pt']

        data = []
        for id in all_samples:
            imgid, sampleid = id.split('_')
            sampleid = ord(sampleid) - 65
            metadata = spectral.envi.open(dataset_path+"/spectral_data/{}-Radiance From Raw Data-Reflectance from Radiance Data and Measured Reference Spectrum.bip.hdr".format(imgid)).metadata
            gt = ast.literal_eval(metadata['gt_TFe'])
            data.append({
                "id": id,
                "gt": torch.Tensor([gt[sampleid]]).to(device)
            })

        # 排序数据集
        data = sorted(data, key=lambda x: x["gt"])
        all_ids = [d["id"] for d in data]
        all_gts = [d["gt"] for d in data]
        max_gt = max(all_gts)
        min_gt = min(all_gts)
        
        # split data into 5 folds, make each fold has the same distribution
        folds = [[],[],[],[],[]]
        for i in range(len(data)):
            folds[i%5].append(data[i]["id"])


        # 加载数据
        data_tensor = test_data(all_ids)

        # 加载模型
        models = [Conv_Diff().to(device) for i in range(5)]
        for i in range(5):
            models[i].load_state_dict(torch.load(pt_path_list[i]))
            models[i].eval()

        nrmse_folds = [[] for i in range(5)]

        for f in range(5):
            for idx in range(data_tensor.__len__()):
                len = data_tensor[idx]["tensor"].shape[0]
                cur = 0
                pixelwise_prediction = []
                while cur + 100 <= len:
                    pixelwise_prediction.append(models[f](data_tensor[idx]["tensor"][cur:cur+100]).to("cpu"))
                    cur += 100

                if cur < len:
                    pixelwise_prediction.append(models[f](data_tensor[idx]["tensor"][cur:]).to("cpu"))

                pixelwise_prediction = torch.cat(pixelwise_prediction, dim=0)
                prediction = pixelwise_prediction.mean()*100
                
                nrmse = torch.sqrt(torch.mean((prediction - data_tensor[idx]["gt"])**2))/(max_gt - min_gt)
                nrmse_folds[f].append(nrmse)

        # 打印结果
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'fold_results'
        sheet['A1'] = '样本编号'
        sheet['B1'] = 'fold1'
        sheet['C1'] = 'fold2'
        sheet['D1'] = 'fold3'
        sheet['E1'] = 'fold4'
        sheet['F1'] = 'fold5'
        sheet['G1'] = '平均值'

        for i in range(2, all_ids.__len__()+2):
            sheet['A'+str(i)] = all_ids[i-2]

            # 判断这个id的样本在哪个fold里是测试集，把对应的格子搞成天蓝色
            for j in range(1, 6):
                if all_ids[i-2] in folds[j-1]:
                    sheet.cell(row=i, column=j+1).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='87CEEB')

            for j in range(2, 7):
                sheet.cell(row=i, column=j).value = nrmse_folds[j-2][i-2].item()

            sheet['G'+str(i)] = '=AVERAGE(B'+str(i)+':F'+str(i)+')'

        wb.save('fold_results2.xlsx')
        print('Done!')
